'''
conda create --name syj_vfss_new python=3.8.5
conda install pytorch torchvision torchaudio cudatoolkit=11.3 -c pytorch -c nvidia

conda install tqdm


conda install pandas

conda install openpyxl

conda install psutil

conda install matplotlib

pip install sklearn

conda install -c conda-forge tensorboard


conda install -c conda-forge torchinfo
'''
import os
import sys
import timeit
from tqdm import tqdm
#import pickle
import psutil
#from sklearn.model_selection import KFold
from sklearn import metrics
#import socket
import conv3dVari_each as conv3dVari
from earlyStopping import EarlyStopping
from returnModel_fixed import returnModel
import autoPath
#from sklearn.model_selection import KFold
from sklearn import metrics
import torch
import torch.utils.data as data_utl
from torch.utils.tensorboard import SummaryWriter
# from torchinfo import summary
# import fixRandom # adaptive average pooling 같은 레이어 있으면 아무 소용 없음
# 데이터 로딩 코드랑, 그 이전 코드랑 분리해서, 반복실험 더 빠르게 하기
def init_train_val(split_file, 
    rgb_root, flow_root, 
    frame_len, 
    sampling, 
    train_batch, 
    num_classes, 
    lr, 
    last_model, 
    modelName, 
    loss_type, 
    fill, pad, 
    bi, tri=False):
    # gray_scale=False

    save_dir = os.path.dirname(os.path.abspath(__file__)) # 현재 파일이 있는 디렉토리

    split_file, splitFileName = autoPath.findFile(split_file, save_dir)
    flow_root = autoPath.findDir(flow_root, save_dir)
    rgb_root = autoPath.findDir(rgb_root, save_dir)
    if rgb_root is not None:
        if flow_root is not None:
            input_type='both'
            sys.exit('supported in VariTrain_both.py')
        else:
            input_type='rgb'
    elif flow_root is not None:
        input_type='flow'
    
    if (loss_type!='CE') and (loss_type!='weightedCE') and (loss_type!='ASL') and ('multi' not in loss_type.lower()) and ('poly' not in loss_type.lower()) and (loss_type.lower()!='custom') and (loss_type!='PECE'):
        sys.exit('check loss_type again')

    resume_epoch=0
    if last_model is not None:
        last_model = torch.load(last_model, map_location=lambda storage, loc: storage)   # Load all tensors onto the CPU
        resume_epoch=last_model['epoch']-1
        
    midLabelModel = ['CNN3D', 'i3d', 'VTN', 'vgg', 'efficient', 'vgg3d','resNet3D', 'resNet3Dbi', 'resNet3D_bi', 'resNet3D_fill16', 'TDN']
    resize, labelType, initial_model, optimizer, scheduler, transform, _, __, ___ = returnModel(modelName, resume_epoch, save_dir, num_classes, lr, input_type, frame_len, bi)
    if '512' in sampling:
        resize=(512, 512)
        sampling = sampling.replace('512', '')

    #모델과 tensorboard graph를 저장할 디렉토리 
    if 'multi' in split_file.lower():
        folderName = 'MultiLabel'
    elif labelType == 'label':
        folderName = 'VariModels_labelPro'
    elif labelType == 'frameLabel':
        folderName = 'VariModels_frameLabel'
    elif labelType == 'IoU':
        folderName = 'VariModels'
    os.makedirs(os.path.join(save_dir, folderName), exist_ok=True)
    if fill is not None:
        modelName = modelName+'_fill'+str(fill)
    elif pad is not None:
        modelName = modelName+'_pad'+str(pad)
    if bi:
        modelName = modelName+'_bi'
        frame_len4dataset = frame_len*2-1
    elif tri:
        modelName = modelName+'_tri'
        frame_len4dataset = frame_len*2-1
    else:
        frame_len4dataset = frame_len
    #모델과 tensorboard graph를 저장할 이름
    saveName = f'{input_type.upper()}model-{modelName}_frame-{frame_len}_sample-{sampling}_split-{splitFileName}_LR-{lr}_batch-{train_batch}_loss-{loss_type}'#_iou-{iouTH}
    log_dir = os.path.join(save_dir, folderName, saveName)
    os.makedirs(log_dir, exist_ok=True)
    
    # load_whole=False
    cfg={'resume_epoch': resume_epoch, 'midLabelModel':midLabelModel, 'resize':resize, 'labelType':labelType, 
    'initial_model':initial_model, 'optimizer':optimizer, 'scheduler':scheduler, 'transform':transform, 'last_model':last_model, 
    'frame_len4dataset':frame_len4dataset, 'log_dir':log_dir, 'split_file':split_file, 'rgb_root':rgb_root, 'flow_root':flow_root}
    return cfg # resume_epoch, midLabelModel, resize, labelType, initial_model, optimizer, scheduler, transform, frame_len4dataset, log_dir

def dataloaders(split_file, 
    rgb_root, flow_root, 
    frame_len4dataset, 
    labelType, 
    fill, 
    sampling, 
    resize, 
    num_classes, 
    transform, 
    gray_scale, 
    train_batch, val_batch):

    if 'multi' in split_file.lower():
        if psutil.virtual_memory().available > 4e+10:
            import vfssData_multiLabel as vfssData
        else: 
            sys.exit('this data loader needs more memory')
    elif train_batch==val_batch==1:
        import trainBy1batch_dataClass as vfssData
    elif ('Skipped' in split_file) or ('interpolated' in split_file):
        # print('it is vfssData_withSkippedOrInterpolated')
        # import vfssData_withSkippedOrInterpolatedPreLoaded_revisedFrameLabel as vfssData

        # if psutil.virtual_memory().available > 4e+10:
        #     import vfssData_withSkippedOrInterpolatedPreLoaded as vfssData
        # else:
        import vfssData_withSkippedOrInterpolatedFixed as vfssData 
    else:
        # print('it is just vfssData')
        import vfssData as vfssData
        sys.exit('it will cause error by labelType parameter when making dataset class')

    print('load train data')
    train_dataset = vfssData.VFSS_data(split_file = split_file, split = 'train', rgb_root=rgb_root, flow_root=flow_root, frame_len=frame_len4dataset, labelType = labelType,
        fill=fill, sampling=sampling, resize=resize, num_classes=num_classes, transforms=transform, load_whole=False, gray_scale=gray_scale)
    num_workers=min(round(len(train_dataset)/train_batch+0.4), psutil.cpu_count()//8)
    # print('train num worker: ', num_workers)
    train_dataloader = torch.utils.data.DataLoader(train_dataset, batch_size=train_batch, shuffle=True, num_workers=num_workers)#,worker_init_fn=fixRandom.seed_worker, generator=g)
    print('load validation data')
    val_dataset = vfssData.VFSS_data(split_file = split_file, split = 'val', rgb_root=rgb_root, flow_root=flow_root, frame_len=frame_len4dataset, labelType = labelType,
        fill=fill, sampling=sampling, resize=resize, num_classes=num_classes, transforms=transform, load_whole=False, gray_scale=gray_scale)
    num_workers=min(round(len(val_dataset)/val_batch+0.4), psutil.cpu_count()//8)
    # print('val num worker: ', num_workers)
    val_dataloader = torch.utils.data.DataLoader(val_dataset, batch_size=val_batch, num_workers=num_workers)#,worker_init_fn=fixRandom.seed_worker, generator=g)
    
    trainval_loaders = {'train': train_dataloader, 'val': val_dataloader}
    if labelType == 'frameLabel':
        trainval_sizes = {x: len(trainval_loaders[x].dataset)*frame_len4dataset for x in ['train', 'val']}
    else:
        trainval_sizes = {x: len(trainval_loaders[x].dataset) for x in ['train', 'val']}
    
    return trainval_loaders, trainval_sizes

def on_device(initial_model, last_model, optimizer, loss_type):
    # Use GPU if available else revert to CPU
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f'use device {device}')
    model=initial_model.to(device)
    if torch.cuda.device_count() > 1:
        model = torch.nn.DataParallel(model)
    # DataParallel로 만든 state dict는 DataParallel로 만든 모델에만 넣어야 한다.
    if last_model is not None:
        model.load_state_dict(last_model['state_dict'])
        optimizer.load_state_dict(last_model['opt_dict'])
    # print(summary(model, input_size=(train_batch, in_channels, frame_len, resize[0], resize[1]), verbose=0))
    print('Total params: %.2fM' % (sum(p.numel() for p in initial_model.parameters()) / 1000000.0))
    print('Initial learning rate: ', optimizer.param_groups[0]['lr'])
    if loss_type=='CE':
        criterion = torch.nn.CrossEntropyLoss().to(device)  # standard crossentropy loss for classification
    elif loss_type=='PECE':
        criterion = torch.nn.CrossEntropyLoss(ignore_index=-1).to(device)  # standard crossentropy loss for classification
    # elif loss_type=='AFL':
    #     criterion = conv3dVari.AsymmetricFocalLoss()
    elif loss_type == 'weightedCE':
        criterion = torch.nn.CrossEntropyLoss(weight = torch.tensor([0.1, 0.5, 0.1, 0.3])).to(device)
    elif loss_type=='ASL': # mean 아니고 sum임.
        criterion = conv3dVari.AsymmetricLoss().to(device)
    elif 'multi' in loss_type.lower():
        criterion = torch.nn.MultiLabelMarginLoss().to(device)
    elif 'poly' in loss_type.lower(): # output과 label 모양이 [batch, num_classes, frame_len]
        # label_is_onehot=True 이면, num_classes정보 안쓰인다.
        # 그런데 weight은 num_classes 수에 맞는 걸 써야 한다.
        # criterion = conv3dVari.Poly1FocalLoss(4, reduction='mean', weight = torch.tensor([0.1, 0.5, 0.1, 0.3], device=device), label_is_onehot=True).to(device)
        # criterion = conv3dVari.Poly1FocalLoss(6, reduction='mean', weight = torch.tensor([0.1, 0.2, 0.1, 0.1, 0.2, 0.3], device=device), label_is_onehot=True).to(device)
        # criterion = conv3dVari.Poly1FocalLoss(3, reduction='mean', weight = torch.tensor([0.3, 0.3, 0.4], device=device), label_is_onehot=True).to(device)
        # criterion = conv3dVari.Poly1FocalLoss(2, reduction='mean', weight = torch.tensor([0.1, 0.9], device=device), label_is_onehot=True).to(device) #binary[negative, positive]
        criterion = conv3dVari.Poly1FocalLoss(4, reduction='mean', weight = torch.tensor([0.25, 0.25, 0.25, 0.25], device=device), label_is_onehot=True).to(device)
        # weight 후보
        # [0.01, 0.165, 0.165, 0.165, 0.165, 0.165, 0.165]
        # [0.04, 0.11, 0.11, 0.11, 0.11, 0.11, 0.11]
        # [0.07, 0.105, 0.105, 0.105, 0.105, 0.105, 0.105]
    elif loss_type.lower()=='custom':#labelType == 'frameLabel' or frame_len==1:
        weight = torch.tensor([1,3])# the loss weight when num_classes==2
        weight = weight/weight.sum()
        criterion = conv3dVari.PenaltyLoss(torch.nn.CrossEntropyLoss(weight=weight)).to(device)

    return model, optimizer, criterion

def Labeling(labelType, labels, bi, tri, modelName, midLabelModel, frame_len, iouLabel=None):
    # multiLabel도 혼용할 수 있게 바꿀 수 있을까
    # if labels.size(1) > 2:
    #     # labels.shape이 (batch, num_classes, frame_len 혹은 feature로 나옴)
    #     label = labels.long()
    if labelType == 'frameLabel': 
        if labels.size(1) > 2:
            label = labels.transpose(1, 2).contiguous().flatten(0, 1).long() # (batch * frame_len, num_classes)
        else:
            label = torch.flatten(labels[:, 1], start_dim=0).long()
        # print('frameLabel label size:', label.size())#torch.Size([frame_len])
    elif bi or tri or (modelName in midLabelModel): # print(labels.size())# ([batch, num_classes=2, frame_len])
        if bi or tri: 
            ind = [frame_len-1] # bi 는 무조건 홀수다!! # frame_len == 5 인 경우, 00001+10000 => 000010000 이므로.
        else:
            if frame_len%2 == 0: # frame_len == 4 인 경우, 0110
                ind = [int(frame_len/2-1), int(frame_len/2-2)]
            else:# frame_len == 5 인 경우, 00100
                ind = [int(frame_len/2)]
            # window 중간이 positive인가?
        # print('label index: ', ind)
        label = (torch.sum(labels[:, 1, ind], 1)>=len(ind)).long() #size([1])
    elif labelType == 'label':
        TH = frame_len/2
        label = (torch.sum(labels[:, 1], 1)>=TH).long()
        # print('3')
    elif labelType == 'IoU':
        label = torch.as_tensor(iouLabel).long()
        # print('4')

    return label

def cal_epoch_loss_acc_ap(running_loss, y_true, y_score, data_len):
    epoch_loss = running_loss / data_len
    epoch_acc = metrics.accuracy_score(y_true, y_score)
    epoch_ap = metrics.average_precision_score(y_true, y_score, average='micro')#AP는 Recall에 따른 Precision 곡선 아래 면적
    #현재 클래스 하나만 detect하는 것이므로, mAP==AP
    return epoch_loss, epoch_acc, epoch_ap

def train_anEpoch(labelType, model, optimizer, scheduler, device, 
    frame_len, 
    trainLoaders, train_size, # trainval_loaders['train'], trainval_sizes['train']
    modelName, midLabelModel, 
    loss_type, criterion, sampling,
    bi, tri=False):
    if tri: # tri 안쓸거라서 없어도 되는데, 코드 통일을 위해 일단 이 함수 안에 필요함.
        start_ind = frame_len - frame_len//2 - 1
        end_ind = frame_len + frame_len//2 - 1
    # reset the running loss and corrects
    running_loss = 0.0
    y_true, y_score = [], []
    model.train()
    for data in tqdm(trainLoaders):
        if 'PE' in sampling:
            inputs, labels, patientID, start_f = data # iouLabel, 
            start_f = start_f.to(device)
            # print(inputs.shape, start_f) #torch.Size([5, 3, 50, 224, 224]) tensor([ 0, 33, 20, 66, 44], device='cuda:0')
        elif (modelName == 'vgg') or (modelName == 'i3d'):
            inputs, labels, iouLabel, patientID = data #  
        else:
            inputs, labels, patientID = data # iouLabel, 
        # move inputs and labels to the device the training is taking place on
        inputs = inputs.to(device)#Size([batch, RGB, frame_len, H, W])
        # if (modelName == 'vgg') or 'efficient' in modelName:
        #     print(inputs.size())
        #     inputs = inputs.permute(0, 2, 1, 3, 4).contiguous()
        #     inputs = torch.flatten(inputs, start_dim=0, end_dim=1)
        #     #Size([batch*frame_len, RGB, H, W])
        # elif modelName=='TDN':# make Size([batch, frame_len, RGB, W, H])
        #     inputs = inputs.permute(0, 2, 1, 4, 3).contiguous()
        optimizer.zero_grad()
        if bi:# frame_len == 5 인 경우, 00001+10000 => 000010000 이므로.
            outputs = model(inputs[:, :, :frame_len], inputs[:, :, frame_len-1:])
        elif tri:
            outputs = model(inputs[:, :, :frame_len], inputs[:, :, start_ind:end_ind], inputs[:, :, frame_len-1:])
        elif sampling == 'PE':
            outputs = model(inputs, start_f)
        elif sampling == 'PECE':
            NP_index = (labels[:, 1]==-1).nonzero(as_tuple=True) # paddig indices
            # print(labels[:, 1].shape) # (batch, frame_len)
            # print(NP_index.shape) # (-1인 위치 수, 차원갯수(행렬=2)) # .nonzero(as_tuple=False) 일 경우.
            # print(NP_index[0].shape, NP_index[1].shape) # (-1인 위치 수), (-1인 위치 수)
            key_padding_mask = torch.full((inputs.size(0), inputs.size(2)), False)
            # print(inputs.shape) # (batch, channel, frame_len, H, W)
            # print(key_padding_mask.shape) # (batch, frame_len)
            key_padding_mask[NP_index] = True
            key_padding_mask = key_padding_mask.to(device)
            outputs = model(inputs, start_f, mask=key_padding_mask)
        else:
            outputs = model(inputs)

        if torch.isnan(torch.sum(outputs)): sys.exit("There is NaN in model's output")

        # if outputs.size(-1) > 2:
        #     outputs = outputs.transpose(1, 2).contiguous() 
        if labelType == 'frameLabel': # outputs: torch.Size([batch, frame_len, num_classes)
            outputs=torch.flatten(outputs, start_dim=0, end_dim=1) # torch.Size([batch * frame_len, num_classes)
        else:
            if ('i3d' in modelName) and (len(outputs.size())==3) and (outputs.size(-1)==1): 
                outputs = outputs.squeeze(2)
            elif modelName=='TDN': outputs = outputs.squeeze(1)
            # print(outputs.size())#torch.Size([batch=10, num_classes=2])                  
        
        # labelType == 'IoU': folderName = 'VariModels' 인 경우는 아래 함수에 iouLabel을 넣어야 함
        label=Labeling(labelType, labels, bi, tri, modelName, midLabelModel, frame_len) #, iouLabel
        label=label.to(device)

        
        if loss_type.lower()=='custom':
            preds = torch.nn.Softmax(dim=1)(outputs) #torch.Size([batch=10, num_classes=2])
            preds = torch.max(preds, 1)[1] #torch.Size([frame_len])
            loss = criterion(outputs, label, preds)
        else:
            if ('CE' in loss_type) and ('multi' in trainLoaders.dataset.split_file.lower()):
                label = label.float()
            loss = criterion(outputs, label) 
            with torch.no_grad():
                if labels.size(1) > 2:
                    preds = torch.nn.Sigmoid()(outputs) # torch.Size([batch * frame_len, num_classes)
                    preds = preds > 0.5
                else:
                    preds = torch.nn.Softmax(dim=1)(outputs) #torch.Size([batch=10, num_classes=2])
                    preds = torch.max(preds, 1)[1] #torch.Size([frame_len])
        
        loss.backward()
        optimizer.step()
        if (scheduler is not None) and ('ReduceLROnPlateau' not in str(scheduler)):
            scheduler.step()

        with torch.no_grad():
            # PECE 의 경우 loss를 계산하거나 y_true, y_score 리스트 생성 시 ignore_index를 하는 레이블 제외하고 하기.
            # multiLabel의 경우, one-hot vector로 만들어놓았는데, padding이 되는 부분을 모든 class에 대해 ignore_index 하면 된다..!
            # print(outputs.shape) # (batch x frame_len, num_classes) or (batch, num_classes)
            # print(label.shape) # (batch) if binary, (batch, num_classes) if multiLabel
            if loss_type=='ASL':
                running_loss += loss.item()
            elif loss_type=='PECE':
                NP_index = (label!=-1).nonzero(as_tuple=True)[0] # non-paddig indices
                running_loss += loss.item() * len(NP_index)
                label = label[NP_index]
                preds = preds[NP_index]
            else:
                running_loss += loss.item() * outputs.size(0)
            
            if labels.size(1) > 2:
                label = label.cpu()
                preds = preds.cpu()
                if y_true == []:
                    y_true = label
                    y_score = preds
                else:
                    y_true = torch.cat((y_true, label), dim=0)
                    y_score = torch.cat((y_score, preds), dim=0)
            else:
                y_true.extend(label.tolist())
                y_score.extend(preds.tolist())
    # model, optimizer, schedular, loss 다 return 해서 eval_anEpoch에 넘겨야 하는거 아닌가?
    # cfg라는 dictionary혹은 yaml파일을 넘기는 게 낫지 않나
    # argparser로 dictionary나 yaml파일 쓸 수 있는 지 알아보기
    return cal_epoch_loss_acc_ap(running_loss, y_true, y_score, train_size)


@torch.no_grad() # with torch.no_grad() 역할 하는 데코레이터!
def eval_anEpoch(labelType, model, scheduler, device, 
    frame_len, 
    valLoaders, val_size, # trainval_loaders['val'], trainval_sizes['val'] # data_len 을 인자로 넘겨주는 코드 짜서 없애기
    # log_dir, writer, #early_stopping
    modelName, midLabelModel, 
    # n_epochs_stop, 
    loss_type, criterion, sampling,
    bi, tri=False):
    if tri: # tri 안쓸거라서 없어도 되는데, 코드 통일을 위해 일단 이 함수 안에 필요함.
        start_ind = frame_len - frame_len//2 - 1
        end_ind = frame_len + frame_len//2 - 1
    # reset the running loss and corrects
    running_loss = 0.0
    y_true, y_score = [], []
    model.eval()
    for data in tqdm(valLoaders):
        if 'PE' in sampling:
            inputs, labels, patientID, start_f = data
            start_f = start_f.to(device)
        elif (modelName == 'vgg') or (modelName == 'i3d'):
            inputs, labels, iouLabel, patientID = data # 
        else:
            inputs, labels, patientID = data
        # move inputs and labels to the device the training is taking place on
        inputs = inputs.to(device)#Size([batch, RGB, frame_len, H, W])
        # if (modelName == 'vgg') or 'efficient' in modelName:
        #     print(inputs.size())
        #     inputs = inputs.permute(0, 2, 1, 3, 4).contiguous()
        #     inputs = torch.flatten(inputs, start_dim=0, end_dim=1)
        #     #Size([batch*frame_len, RGB, H, W])
        # elif modelName=='TDN':# make Size([batch, frame_len, RGB, W, H])
        #     inputs = inputs.permute(0, 2, 1, 4, 3).contiguous()
        with torch.no_grad():
            if bi:
                outputs = model(inputs[:, :, :frame_len], inputs[:, :, frame_len-1:])
            elif tri:
                outputs = model(inputs[:, :, :frame_len], inputs[:, :, start_ind:end_ind], inputs[:, :, frame_len-1:])
            elif sampling == 'PE':
                outputs = model(inputs, start_f)
            elif sampling == 'PECE':
                NP_index = (labels[:, 1]==-1).nonzero(as_tuple=True) # non-paddig indices
                key_padding_mask = torch.full((inputs.size(0), inputs.size(2)), False)
                key_padding_mask[NP_index] = True
                key_padding_mask = key_padding_mask.to(device)
                outputs = model(inputs, start_f, mask=key_padding_mask)
            else:
                outputs = model(inputs)

        if torch.isnan(torch.sum(outputs)): sys.exit("There is NaN in model's output")
    
        if labelType == 'frameLabel': # outputs: torch.Size([batch, frame_len, num_classes)
            outputs=torch.flatten(outputs, start_dim=0, end_dim=1)
        else:
            if ('i3d' in modelName) and (len(outputs.size())==3) and (outputs.size(-1)==1): 
                outputs = outputs.squeeze(2)
            elif modelName=='TDN': outputs = outputs.squeeze(1)
            # print(outputs.size())#torch.Size([batch=10, num_classes=2])  
        
        # labelType == 'IoU': folderName = 'VariModels' 인 경우는 아해 함수에 iouLabel을 넣어야 함
        label=Labeling(labelType, labels, bi, tri, modelName, midLabelModel, frame_len) # , iouLabel
        label=label.to(device)

        with torch.no_grad():
            if loss_type.lower()=='custom':
                preds = torch.nn.Softmax(dim=1)(outputs) #torch.Size([batch=10, num_classes=2])
                preds = torch.max(preds, 1)[1] #torch.Size([frame_len])
                loss = criterion(outputs, label, preds)
            else:
                if ('CE' in loss_type) and ('multi' in valLoaders.dataset.split_file.lower()):
                    label = label.float()
                loss = criterion(outputs, label)
                if labels.size(1) > 2:
                    preds = torch.nn.Sigmoid()(outputs) # torch.Size([batch * frame_len, num_classes)
                    preds = preds > 0.5
                else:
                    preds = torch.nn.Softmax(dim=1)(outputs) #torch.Size([batch=10, num_classes=2])
                    preds = torch.max(preds, 1)[1] #torch.Size([frame_len])
                
            if loss_type=='ASL':
                running_loss += loss.item()
            elif loss_type=='PECE':
                NP_index = (label!=-1).nonzero(as_tuple=True)[0] # non-paddig indices
                running_loss += loss.item() * len(NP_index)
                label = label[NP_index]
                preds = preds[NP_index]
                val_size = len(NP_index)
            else:
                running_loss += loss.item() * outputs.size(0)
            
            if labels.size(1) > 2:
                label = label.cpu()
                preds = preds.cpu()
                if y_true == []:
                    y_true = label
                    y_score = preds
                else:
                    y_true = torch.cat((y_true, label), dim=0)
                    y_score = torch.cat((y_score, preds), dim=0)
            else:
                y_true.extend(label.tolist())
                y_score.extend(preds.tolist())

    epoch_loss, epoch_acc, epoch_ap = cal_epoch_loss_acc_ap(running_loss, y_true, y_score, val_size)
    if 'ReduceLROnPlateau' in str(scheduler):
        scheduler.step(epoch_loss)

    return epoch_loss, epoch_acc, epoch_ap, y_true, y_score

def draw_on_tensorboard(y_true, y_score, 
    writer, 
    modelName, 
    train_loss, train_acc, train_ap, 
    epoch_loss, epoch_acc, ap, 
    epoch, 
    optimizer,
    num_classes):

    if num_classes > 2:
        auc = metrics.roc_auc_score(y_true, y_score,  average='micro')
    else:
        fpr, tpr, thresholds = metrics.roc_curve(y_true, y_score, pos_label=1)
        auc = metrics.auc(fpr, tpr)
    recall = metrics.recall_score(y_true, y_score, pos_label=1, average='micro') # TP / (TP + FN)
    precision = metrics.precision_score(y_true, y_score, pos_label=1, average='micro') # TP / TPFP
    F1_score = metrics.f1_score(y_true, y_score, pos_label=1, average='micro') # 2*precision*recall/(precision+recall)
    # average option에서 'samples'는 멀티레이블만 가능, 
    # 'binary'는 바이너리 데이터만 가능하고, 바이너리 데이터가 들어가면 다른 average option을 넣어도 'binary' option과 똑같이 나온다, 
    # None은 각 클래스 비교한 갯수별로 스코어 나옴. 바이너리 데이터면 하나 나옴
    # 대부분(다른 논문들)은 sum을 하는 것으로 보아 'micro'를 사용하는 것 같음. 뭐로 나누거나 하는 것을 못봣음
    # FNR = FN/(FN+TP)#False Negative Rate
    if num_classes > 2:
        fp = torch.logical_and(torch.logical_not(y_true), y_score).float().mean()
    else:
        fp = torch.logical_and(torch.logical_not(torch.tensor(y_true)), torch.tensor(y_score)).float().mean()

    # multilabel용 make label array 만드는 함수 결과 matrix 받고 거기 있는 레이블별로 딕셔너리 만들게 하기. val 값으로만!!
    # dataLen = len(y_true)
    # for label in labelArray:
    #   writer.add_scalars(f'{modelName}/{label}', {'true': y_true.count(label) / dataLen, 'pred': y_score.count(label) / dataLen}, epoch)
    # zero_ratio = y_score.count(0)/len(y_score) # TNFN / (data_len)
    # one_ratio = y_score.count(1)/len(y_score) # TPFP / (data_len)
    print("recall: {}".format(round(recall.item(), 3)))
    print("precision: {}".format(round(precision.item(), 3)))
    print("F1_score: {}".format(round(F1_score.item(), 3)))
    print("AUC: {}".format(round(auc, 3)))
    print("AP: {}".format(round(ap, 3)))
    # print("zero_ratio: {}, one_ratio: {}".format(zero_ratio, one_ratio))
    writer.add_scalars(f'{modelName}/loss', {'train': train_loss, 'val': epoch_loss}, epoch)
    writer.add_scalars(f'{modelName}/accuracy', {'train': train_acc, 'val': epoch_acc}, epoch)
    writer.add_scalars(f'{modelName}/average_precision', {'train': train_ap, 'val': ap}, epoch)
    writer.add_scalar(f'{modelName}/learning_rate', optimizer.param_groups[0]['lr'], epoch)
    writer.add_scalar(f'{modelName}_val/recall', recall, epoch)
    writer.add_scalar(f'{modelName}_val/precision',precision, epoch)
    writer.add_scalar(f'{modelName}_val/F1_score', F1_score, epoch)
    writer.add_scalar(f'{modelName}_val/AUC', auc, epoch)
    for i in range(num_classes): # one hot multi
        if str(type(y_score)) == "<class 'list'>":
            f1 = metrics.f1_score(y_true, y_score) # multi class is not available
        else: # if str(type(y_score)) == "<class 'torch.Tensor'>"
            f1 = metrics.f1_score(y_true[:, i], y_score[:, i])

        writer.add_scalar(f'{modelName}_val/class_{i}_F1', f1, epoch)
    # writer.add_scalar(f'{modelName}_val/output_one_ratio', one_ratio, epoch)
    # for i in range(num_classes): # one hot multi
    #     if str(type(y_score)) == "<class 'list'>":
    #         ratio = y_score.count(1) / len(y_score) # multi class is not available
    #     else: # if str(type(y_score)) == "<class 'torch.Tensor'>"
    #         ratio = torch.count_nonzero(y_score[:, i]) / y_score.size(0)
            
    #     writer.add_scalar(f'{modelName}_val/output_{i}_ratio', ratio, epoch)
    writer.add_scalar(f'{modelName}_val/False_Positive', fp, epoch)

def train_val(labelType, model, optimizer, scheduler, 
    epoch, num_epochs, 
    frame_len, device, 
    trainval_loaders, trainval_sizes, 
    writer, #log_dir, 
    modelName, midLabelModel, 
    # n_epochs_stop, 
    loss_type, criterion, sampling,
    num_classes,
    bi, tri=False):

    for phase in ['train', 'val']:
        start_time = timeit.default_timer()
        if phase == 'train':  
            train_loss, train_acc, train_ap = train_anEpoch(labelType=labelType, model=model, 
                                                            optimizer=optimizer, scheduler=scheduler, 
                                                            device=device, frame_len=frame_len, 
                                                            modelName=modelName, midLabelModel=midLabelModel,
                                                            loss_type=loss_type, criterion=criterion, sampling = sampling,
                                                            bi=bi, tri=tri,
                                                            trainLoaders=trainval_loaders[phase], 
                                                            train_size=trainval_sizes[phase])
            epoch_loss=train_loss
            epoch_acc=train_acc
        else:
            epoch_loss, epoch_acc, epoch_ap, y_true, y_score = eval_anEpoch(labelType=labelType, model=model, 
                                                                            scheduler=scheduler, device=device, 
                                                                            frame_len=frame_len, modelName=modelName,
                                                                            midLabelModel=midLabelModel, loss_type=loss_type,
                                                                            criterion=criterion, sampling = sampling, 
                                                                            bi=bi, tri=tri,
                                                                            valLoaders=trainval_loaders[phase], 
                                                                            val_size=trainval_sizes[phase])
                                                                            # n_epochs_stop, log_dir, writer, #early_stopping
        print("[{}] Epoch: {}/{} Loss: {} Acc: {}".format(phase, epoch+1, num_epochs, epoch_loss, epoch_acc))
        stop_time = timeit.default_timer()
        print("Execution time: " + str(stop_time - start_time) + "\n")

    draw_on_tensorboard(y_true=y_true, y_score=y_score, writer=writer, modelName=modelName, 
                        train_loss=train_loss, train_acc=train_acc, train_ap=train_ap, 
                        epoch_loss=epoch_loss, epoch_acc=epoch_acc, ap=epoch_ap, 
                        epoch=epoch, optimizer=optimizer, num_classes=num_classes)
    return epoch_loss

def trainVal4epochs(log_dir, n_epochs_stop, resume_epoch, num_epochs, 
                labelType, model, optimizer, scheduler,
                frame_len, trainval_loaders, trainval_sizes, 
                modelName, midLabelModel, loss_type, criterion, sampling,
                num_classes, bi, tri=False):

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    writer = SummaryWriter(log_dir=log_dir)
    if n_epochs_stop is not None: early_stopping = EarlyStopping(patience = n_epochs_stop, verbose = False, path=log_dir+'.pth.tar')
    
    for epoch in range(resume_epoch, num_epochs):
        epoch_loss=train_val(labelType=labelType, 
                    model=model, optimizer=optimizer, scheduler=scheduler, 
                    epoch=epoch, num_epochs=num_epochs, 
                    frame_len=frame_len, device=device, 
                    trainval_loaders=trainval_loaders, trainval_sizes=trainval_sizes, 
                    writer=writer, modelName=modelName, midLabelModel=midLabelModel, 
                    loss_type=loss_type, criterion=criterion, sampling=sampling,
                    num_classes=num_classes, bi=bi, tri=tri)
        if n_epochs_stop is not None:
            early_stopping(epoch_loss, model, epoch, optimizer)
            if early_stopping.early_stop:
                model4test = log_dir+'.pth.tar'
                break
    if early_stopping.early_stop: writer.close()

    return model4test

import numpy as np
import matplotlib.pyplot as plt
import openpyxl
# vfssRGBsingleEval_anEpoch_PE1stF.py 파일과 동일
@torch.no_grad() # with torch.no_grad() 역할 하는 데코레이터!
def loadWhole_eval(test_dataloader, model, labelType, front_rear_pad, frame_len, modelName, device, fill, bi, win_limit, criterion, inferroot=None):#, split_file, output_type
    
    if inferroot is not None:
        wb = openpyxl.Workbook()
        sheet = wb[wb.active.title]
        sheet.append(['patientID', 'start', 'end']) # interpolated 가져올때는 interpolation rate 곱해줘야 함
        saveName = inferroot.split("/")[-1]
        resultFile = f'{inferroot}/{saveName.split("_stride")[0]}.xlsx'
        
    model.eval() # # Primarily affects layers such as BatchNorm or Dropout.
    if fill is not None:
        tile_num = fill//frame_len
        partial_num = fill%frame_len
    y_true, y_score, loss  = [], [], 0.0
    for data in tqdm(test_dataloader):
        if 'PE' in test_dataloader.dataset.sampling:
            if (win_limit==1) and ('multi' not in test_dataloader.dataset.split_file.lower()):
                inputs, labels, patientID, start_f_list = data
            elif (modelName == 'vgg') or (modelName == 'i3d'):
                inputs, labels, patientID = data # , iouLabel
            else:
                inputs, labels, patientID, start_f_list = data # , iouLabel
            start_f_list = start_f_list[0]
            # print(start_f_list.size())
        else:
            inputs, labels, iouLabel, patientID = data
        inputs=inputs[0] # (view_num, C, frame_len4dataset, H x W)
        patientID = patientID.item()
        labels = labels[0] # (num_classes, video_length)
        if labels.size(0) > 2:
            multi=True
            labels = labels.transpose(0, 1).contiguous().long()

        else:
            multi=False
            labels = labels[1].long() # (video_length)
        # idx_l = list(range(len(iou_l))) 
        # print('output from data loader: inputs-',inputs.shape, ', label-', labels.shape)
        loop = inputs.size(0) // win_limit
        if inputs.size(0) % win_limit: loop = loop+1
        # print(inputs.shape, labels.shape) #torch.Size([6, 3, 13, 224, 224]) torch.Size([66])
        prob_l = None
        for i in range(loop):
            
            if 'PE' in test_dataloader.dataset.sampling:
                if (win_limit==1) and ('multi' not in test_dataloader.dataset.split_file.lower()):
                    start_f = start_f_list.unsqueeze(0)
                    sub_input = inputs
                else:
                    start_f = start_f_list[i*win_limit:min(inputs.size(0), i*win_limit+win_limit)]
                    sub_input = inputs[i*win_limit:min(inputs.size(0), i*win_limit+win_limit)]
                start_f = start_f.to(device)
                # print(start_f.size())
            else:
                sub_input = inputs[i*win_limit:min(inputs.size(0), i*win_limit+win_limit)]

            if labelType == 'frameLabel':
                sub_label = labels[i*win_limit*frame_len:min(labels.size(0), i*win_limit*frame_len+win_limit*frame_len)].to(device)
            else:
                sub_label = labels[i*win_limit:min(inputs.size(0), i*win_limit+win_limit)].to(device)

            # print(f'sub input shape: {sub_input.shape}')
            # print(f'sub input segment indices: {i*win_limit} to {min(inputs.size(0), i*win_limit+win_limit)}')
            # print(f'sub label shape: {sub_label.shape}')
            # print(f'label segment indices: {i*win_limit*frame_len} to {min(labels.size(0), i*win_limit*frame_len+win_limit*frame_len)}')
            if len(sub_input.size()) == 4: sub_input = sub_input.unsqueeze(0) # 배치가 1인경우 배치 생성
            # print(sub_input.shape) # (stack_num, channel, frame_len, H, W)
            if ('vgg' in modelName) and (frame_len == 1): sub_input = sub_input.squeeze(2)
            elif fill is not None:
                sub_input = sub_input.numpy()
                sub_input = np.transpose(sub_input, (2,1,0,3,4))
                sub_input = np.tile(sub_input, (tile_num, 1, 1, 1, 1))
                if partial_num != 0: sub_input=np.concatenate((sub_input, sub_input[:partial_num]))
                sub_input = np.transpose(sub_input, (2, 1, 0, 3, 4))
                sub_input = torch.from_numpy(sub_input)
            sub_input = sub_input.to(device)
            # print(sub_input.size())
            with torch.no_grad(): # test니까 gradient안흐르게 model에 넣어보기. 아직 TDN 구현 안됨.
                if bi:# 0110, 00100
                    frame_prob = model(sub_input[:, :, :frame_len], sub_input[:, :, frame_len-1:])
                elif test_dataloader.dataset.sampling == 'PE':
                    frame_prob = model(sub_input, start_f)
                elif test_dataloader.dataset.sampling == 'PECE':
                    # print(sub_label.shape) # (video_length, num_classes) if multiLabel else (video_length)
                    # print(sub_input.shape) # (stack_num, channel, frame_len, H, W) -> [batch, frameLength, embed_dim]
                    NP_index = (sub_label==-1).nonzero(as_tuple=True) # paddig indices
                    if multi: NP_index = NP_index[0]
                    key_padding_mask = torch.full((sub_input.size(0), sub_input.size(2)), False) # (stack_num, frame_len)
                    key_padding_mask[NP_index] = True
                    key_padding_mask = key_padding_mask.to(device)
                    frame_prob = model(sub_input, start_f, mask=key_padding_mask)
                else:
                    frame_prob = model(sub_input)
                # print(frame_prob.size())
                if labelType == 'frameLabel':
                    frame_prob = torch.flatten(frame_prob, start_dim=0, end_dim=1)#torch.Size([(batch) x (frame_len), num_classes=2])
                    if front_rear_pad != 0:
                        frame_prob = frame_prob[:-front_rear_pad]
                if modelName=='TDN': frame_prob = frame_prob.squeeze(1)
                # print(frame_prob)
                # print('before cutting: ', frame_prob.shape)
                frame_prob = frame_prob[:len(sub_label)]
                # print('after cutting: ', frame_prob.shape)
                if criterion is not None:
                    if str(criterion)=='AsymmetricLoss()':
                        loss = loss + criterion(frame_prob, sub_label).item()
                    elif test_dataloader.dataset.sampling == 'PECE':
                        NP_index = (sub_label!=-1).nonzero(as_tuple=True)[0] # non-paddig indices
                        running_loss += loss.item() * len(NP_index)
                    else:
                        loss = loss + criterion(frame_prob, sub_label).item()*frame_prob.size(0) # make it 'reduction = sum'
                
                if multi:
                    frame_prob = torch.nn.Sigmoid()(frame_prob)
                    frame_prob = frame_prob > 0.5
                else:                
                    frame_prob = torch.nn.Softmax(dim=1)(frame_prob)# size: (stack_num, num_classes)
                    frame_prob = torch.max(frame_prob, 1)[1]#.squeeze()# size: (stack_num)

            if test_dataloader.dataset.sampling == 'PECE':
                NP_index = (sub_label!=-1).nonzero(as_tuple=True)[0] # non-paddig indices
                sub_label = sub_label[NP_index]
                frame_prob = frame_prob[NP_index]
            
            if multi:
                if prob_l is None:
                    prob_l = frame_prob.cpu()
                else:
                    prob_l = torch.cat((prob_l, frame_prob.cpu()), dim=0)
            else:            
                if prob_l is None: 
                    prob_l = frame_prob.tolist()
                elif type(frame_prob.tolist())==int:
                    prob_l.append(frame_prob.item())
                else: prob_l.extend(frame_prob.tolist())
                # print('predicted length', len(prob_l))
        if multi:
            if y_true == []:
                y_true = labels
                y_score = prob_l
            else:
                y_true = torch.cat((y_true, labels), dim=0)
                y_score = torch.cat((y_score, prob_l), dim=0)
            if inferroot is not None:
                plot_prob_l = prob_l[:len(labels)] # cut off paddding part...used when labelType frameLabel
                idx_l = list(range(len(labels))) 
                # plotting 코드 짜야 됨!!!!!!!!!!!!!!!!!!!!!!!!!
                num_class = labels.size(1)

                for class_no in range(num_class): # torch.Size([frame_len, num_classes])
                    plt.subplot(num_class, 1, class_no+1) # nrows=num_class, ncols=1, index=class_no
                    plt.plot(idx_l, labels[:, class_no], label='GT')
                    plt.plot(idx_l, plot_prob_l[:, class_no], label='Pred')
                    plt.ylabel(f'class {class_no}')
                    plt.legend()
                plt.savefig(inferroot+f'/{patientID}.png')
                plt.clf()
        else:
            y_true.extend(labels)
            y_score.extend(prob_l)
            # print(len(y_true), len(y_score))
            if inferroot is not None:
                plot_prob_l = prob_l[:len(labels)] # cut off paddding part...used when labelType frameLabel
                idx_l = list(range(len(labels))) 
                arg_labels = torch.argwhere(torch.tensor(labels))
                args_prob_l = torch.argwhere(torch.tensor(plot_prob_l))
                plt.plot(idx_l, labels, label=f'GT {min(arg_labels).item()} : {max(arg_labels).item()}')
                if args_prob_l != []:
                    plt.plot(idx_l, plot_prob_l, label=f'Pred {min(args_prob_l).item()} : {max(args_prob_l).item()}')
                else:
                    plt.plot(idx_l, plot_prob_l, label='Pred no' )
                plt.plot(idx_l, labels, idx_l, plot_prob_l)
                plt.savefig(inferroot+f'/{patientID}.png')
                plt.clf()
                indices = [i for i, e in enumerate(plot_prob_l) if e == 1]
                if (indices != []) and sheet is not None: sheet.append([patientID, indices[0], indices[-1]])
        
    if inferroot is not None: wb.save(resultFile)
    if criterion is not None: loss = loss/len(y_true)
    return y_true, y_score, loss 


# vfssRGBsingleTest_variMem_PE1stF.py
def test_model_loadWhole(last_model, rgb_root, flow_root, split_file, num_classes, win_limit):
    if (win_limit==1) and ('multi' not in split_file.lower()):
        from trainBy1batch_dataClass import VFSS_data
    elif ('Skipped' in split_file) or ('interpolated' in split_file) or ('Multi' in split_file) :
        from vfssData_loadWhole_withSkippedOrInterpolated import VFSS_data
    else:
        from vfssData_loadWhole import VFSS_data
        sys.exit('have not tested yet')

    save_dir = os.path.dirname(os.path.abspath(__file__)) # 현재 파일이 있는 디렉토리
    parentdir = os.path.dirname(save_dir)

    split_file, splitFileName = autoPath.findFile(split_file, save_dir)
    if splitFileName not in last_model:
        sys.exit('model and split file mismatch')
    flow_root = autoPath.findDir(flow_root, save_dir)
    rgb_root = autoPath.findDir(rgb_root, save_dir)
    if rgb_root is not None:
        if flow_root is not None:
            input_type='both'
            sys.exit('supported in VariTrain_both.py')
        else:
            input_type='rgb'
            in_channels=3
            root = rgb_root ############# 지울 수 있도록 하기!!!!!!!!!!
    elif flow_root is not None:
        input_type='flow'
        in_channels=2
        root = flow_root ############# 지울 수 있도록 하기!!!!!!!!!!
    else:
        sys.exit('need data root directory')

    last_model_split = last_model.split('/')
    #last_model은 항상 모델이 있는 폴더명부터 줘야 함
    if 'label' in last_model_split[-2]:
        labelType='label'
        folderName = 'VariModels_labelPro'
    elif 'frameL' in last_model_split[-2]:
        labelType='frameLabel'
        folderName='VariModels_frameLabel'
    elif 'Multi' in last_model_split[-2]:
        labelType='frameLabel'
        folderName='MultiLabel'
    else:
        sys.exit('not using now')
        # labelType='IoU'
        # folderName = 'VariModels'

    saveName = last_model_split[-1][:-8]# .pth.tar 제거
    saveNameList = saveName.split('_')
    bi=False
    fill=None
    for e in saveNameList:
        if 'model' in e:
            modelName = e
        elif 'frame' in e:
            frame_len = int(e.split('-')[-1])
        elif 'sample' in e:
            sampling = e.split('-')[-1]
        elif 'LR' in e:
            LR = e[3:]
        elif 'batch' in e:
            batch = int(e.split('-')[-1])
            if batch>win_limit: win_limit=batch
        elif e == 'bi':
            bi = True
        elif 'fill' in e:
            fill = int(e[4:])
    if fill is not None:
        tile_num = fill//frame_len
        partial_num = fill%frame_len
        # elif 'valAcc' in e:
        #     val_acc = e.split('-')[-1]

    # Use GPU if available else revert to CPU
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    resume_epoch=0
    resize, labelType, model, _, __, transform, front_rear_pad, win_stride, target_layer = returnModel(modelName, resume_epoch, save_dir, num_classes, float(LR), input_type, frame_len, bi, trainVal=False)
    model=model.to(device)
    print('Total params: %.2fM' % (sum(p.numel() for p in model.parameters()) / 1000000.0))
    saveName = saveName+'_stride'+str(win_stride)
    inferroot = os.path.join(save_dir, folderName, saveName)
    os.makedirs(inferroot, exist_ok=True)
    # os.makedirs(inferroot+'/grad_cam', exist_ok=True)# VariTest_gradCAM.py 섞기

    checkpoint = torch.load(last_model, map_location=lambda storage, loc: storage)   # Load all tensors onto the CPU
    print("Initializing weights from: {}...".format(last_model)) 
    if 'state_dict' in checkpoint.keys():
        if 'module.' in list(checkpoint['state_dict'].keys())[0]:
            from collections import OrderedDict
            checkpoint['state_dict'] = OrderedDict((k[7:], v) for k, v in checkpoint['state_dict'].items())
        model.load_state_dict(checkpoint['state_dict'])
    else: 
        model.load_state_dict(checkpoint)

    if bi:
        frame_len4dataset = frame_len*2-1
    else:
        frame_len4dataset = frame_len
    test_dataset = VFSS_data(split_file = split_file, split = 'test', rgb_root=rgb_root, flow_root=flow_root, frame_len=frame_len4dataset, win_stride=win_stride,
        fill=fill, sampling=sampling, resize=resize, num_classes=num_classes, transforms=transform, front_rear_pad=front_rear_pad) # , gray_scale=gray_scale
    test_dataloader = torch.utils.data.DataLoader(test_dataset, batch_size=1, shuffle=False, pin_memory=False)    
    # print('dataloader length:',len(test_dataset))
    

    y_true, y_score, _ = loadWhole_eval(test_dataloader, model, labelType, front_rear_pad, frame_len, modelName, device, fill, bi, win_limit, criterion=None, inferroot=inferroot)

    if num_classes > 2:
        auc = metrics.roc_auc_score(y_true, y_score,  average='micro')
    else:
        fpr, tpr, thresholds = metrics.roc_curve(y_true, y_score, pos_label=1)
        auc = metrics.auc(fpr, tpr)
    ap = metrics.average_precision_score(y_true, y_score, average='micro')#AP는 Recall에 따른 Precision 곡선 아래 면적
    accuracy = metrics.accuracy_score(y_true, y_score)
    recall = metrics.recall_score(y_true, y_score, pos_label=1, average='micro') # TP / (TP + FN)
    precision = metrics.precision_score(y_true, y_score, pos_label=1, average='micro') # TP / TPFP
    F1_score = metrics.f1_score(y_true, y_score, pos_label=1, average='micro') # 2*precision*recall/(precision+recall)
    # average option에서 'samples'는 멀티레이블만 가능, 
    # 'binary'는 바이너리 데이터만 가능하고, 바이너리 데이터가 들어가면 다른 average option을 넣어도 'binary' option과 똑같이 나온다, 
    # None은 각 클래스 비교한 갯수별로 스코어 나옴. 바이너리 데이터면 하나 나옴
    # 대부분(다른 논문들)은 sum을 하는 것으로 보아 'micro'를 사용하는 것 같음. 뭐로 나누거나 하는 것을 못봣음
    # FNR = FN/(FN+TP)#False Negative Rate
    
    # multilabel용 make label array 만드는 함수 결과 matrix 받고 거기 있는 레이블별로 딕셔너리 만들게 하기. val 값으로만!!
    # dataLen = len(y_true)
    # for label in labelArray:
    #   writer.add_scalars(f'{modelName}/{label}', {'true': y_true.count(label) / dataLen, 'pred': y_score.count(label) / dataLen}, epoch)
    # zero_ratio = y_score.count(0)/len(y_score) # TNFN / (data_len)
    # one_ratio = y_score.count(1)/len(y_score) # TPFP / (data_len)
    print("Acc: {}".format(round(accuracy.item(), 3)))
    print("recall: {}".format(round(recall.item(), 3)))
    print("precision: {}".format(round(precision.item(), 3)))
    print("F1_score: {}".format(round(F1_score.item(), 3)))
    print("AUC: {}".format(round(auc, 3)))
    print("AP: {}".format(round(ap, 3)))

    if num_classes > 2:
        fp = torch.logical_and(torch.logical_not(y_true), y_score).float().mean()
        print("False Positive: {}".format(round(fp.item(), 3)))
        resultFile = f'multiLabel{num_classes}_output.xlsx'
        if not os.path.isfile(os.path.join(save_dir, resultFile)):
            wb = openpyxl.Workbook()
            sheet = wb[wb.active.title]
            row = ['Test_acc', 'F1_score_total']
            for i in range(num_classes):
                row.append(f'class{i}F1')
            row.extend(['AP', 'model'])
            sheet.append(row)
            wb.save(os.path.join(save_dir, resultFile))

        row = [accuracy, F1_score]
        for i in range(num_classes):
            f1 = metrics.f1_score(y_true[:, i], y_score[:, i])
            row.append(f1)
            print(f"class {i} f1_score: {round(f1.item(), 3)}")
        row.extend([ap, saveName])

        wb = openpyxl.load_workbook(resultFile)
        sheet = wb[wb.active.title] # Sheet1, Sheet
        sheet.append(row)
        wb.save(resultFile) 

    else:
        f1 = metrics.f1_score(y_true, y_score)
        print(f"class 1 f1_score: {round(f1.item(), 3)}")
        resultFile = 'output.xlsx'
        if not os.path.isfile(os.path.join(save_dir, resultFile)):
            wb = openpyxl.Workbook()
            sheet = wb[wb.active.title]
            sheet.append(['LR', 'Test_acc', 'recall', 'precision', 'F1_score', 'AUC', 'AP', 'model'])
            wb.save(os.path.join(save_dir, resultFile))

        wb = openpyxl.load_workbook(resultFile)
        sheet = wb[wb.active.title] # Sheet1, Sheet
        sheet.append([LR, accuracy, recall, precision, F1_score, auc, ap, saveName])
        wb.save(resultFile) 

    # plt.plot(fpr, tpr)
    # plt.savefig(inferroot+f'/ROC_curve_testAcc-{accuracy}_recall-{recall}_precision-{precision}_F1score-{F1_score}_AUC-{auc}_AP-{ap}.png')
    # plt.clf()
    print('the model was: ', last_model)#.split('/')[-1])