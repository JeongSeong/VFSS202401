'''
conda create --name syj_vfss_new python=3.8.5

conda install pytorch torchvision torchaudio cudatoolkit=11.3 -c pytorch -c nvidia

conda install tqdm

conda install pandas

conda install openpyxl

conda install matplotlib

pip install sklearn

conda install scikit-image

conda install -c conda-forge opencv

pip install grad-cam

pip install ttach
'''
# vfssRGBsingleTest.py보다, vfssRGBsingleTestBatchSize.py보다 CPU랑 Meeory 많이 잡아먹지만, 수행속도는 훨씬 빠르다
import os
import autoPath
from returnModel_fixed import returnModel
from sklearn import metrics
import matplotlib.pyplot as plt
import openpyxl
from skimage.util import view_as_windows
import torch
from tqdm import tqdm
import sys
import numpy as np
# VariTest_gradCAM.py 섞기
# from PIL import Image
import timeit
def test_model(last_model, rgb_root, flow_root, split_file, num_classes, win_limit):
    if ('Skipped' in split_file) or ('interpolated' in split_file) :
        from vfssData_loadWhole_withSkippedOrInterpolated import VFSS_data
    else:
        from vfssData_loadWhole import VFSS_data
        sys.exit('have not tested yet')

    # # VariTest_gradCAM.py 섞기 -- 이 부분은 필요없을듯
    # from torchvision.transforms import Grayscale
    # import cv2
    # import numpy as np
    # from pytorch_grad_cam import GradCAM
    # from pytorch_grad_cam.utils.image import show_cam_on_image

    save_dir = os.path.dirname(os.path.abspath(__file__)) # 현재 파일이 있는 디렉토리
    parentdir = os.path.dirname(save_dir)

    split_file, splitFileName = autoPath.findFile(split_file, save_dir)
    if splitFileName not in last_model:
        sys.exit('model and split file mismatch')
    flow_root = autoPath.findDir(flow_root, save_dir)
    rgb_root = autoPath.findDir(rgb_root, save_dir)
    if rgb_root is not None:
        if flow_root is not None:
            input_type='both'
            sys.exit('supported in VariTrain_both.py')
        else:
            input_type='rgb'
            in_channels=3
            root = rgb_root ############# 지울 수 있도록 하기!!!!!!!!!!
    elif flow_root is not None:
        input_type='flow'
        in_channels=2
        root = flow_root ############# 지울 수 있도록 하기!!!!!!!!!!
    else:
        sys.exit('need data root directory')

    last_model_split = last_model.split('/')
    #last_model은 항상 모델이 있는 폴더명부터 줘야 함
    if 'label' in last_model_split[-2]:
        labelType='label'
        folderName = 'VariModels_labelPro'
    elif 'frameL' in last_model_split[-2]:
        labelType='frameLabel'
        folderName='VariModels_frameLabel'
    else:
        sys.exit('not using now')
        # labelType='IoU'
        # folderName = 'VariModels'

    saveName = last_model_split[-1][:-8]# .pth.tar 제거
    saveNameList = saveName.split('_')
    bi=False
    fill=None
    for e in saveNameList:
        if 'model' in e:
            modelName = e
        elif 'frame' in e:
            frame_len = int(e.split('-')[-1])
        elif 'sample' in e:
            sampling = e.split('-')[-1]
        elif 'LR' in e:
            LR = e[3:]
        elif 'batch' in e:
            batch = int(e.split('-')[-1])
            # if batch>win_limit: win_limit=batch
        elif e == 'bi':
            bi = True
            # sys.exit('gradCAM only takes a tensor as input_tensor. If you want to run this code without yielding gradCAM, remove this cod line, then run this code.')
        elif 'fill' in e:
            fill = int(e[4:])
    if fill is not None:
        tile_num = fill//frame_len
        partial_num = fill%frame_len
        # elif 'valAcc' in e:
        #     val_acc = e.split('-')[-1]

    # Use GPU if available else revert to CPU
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    resume_epoch=0
    resize, labelType, model, _, __, transform, front_rear_pad, win_stride, target_layer = returnModel(modelName, resume_epoch, save_dir, num_classes, float(LR), input_type, frame_len, bi, trainVal=False)
    if '512' in sampling:
        resize = (512, 512)
        sampling = sampling.replace('512', '')

    model=model.to(device)
    print('Total params: %.2fM' % (sum(p.numel() for p in model.parameters()) / 1000000.0))
    saveName = saveName+'_stride'+str(win_stride)
    inferroot = os.path.join(save_dir, folderName, saveName)
    os.makedirs(inferroot, exist_ok=True)
    # if (labelType != 'frameLabel') and not bi:
    #     os.makedirs(inferroot+'/grad_cam', exist_ok=True)# VariTest_gradCAM.py 섞기

    checkpoint = torch.load(last_model, map_location=lambda storage, loc: storage)   # Load all tensors onto the CPU
    print("Initializing weights from: {}...".format(last_model)) 
    if 'state_dict' in checkpoint.keys():
        if 'module.' in list(checkpoint['state_dict'].keys())[0]:
            from collections import OrderedDict
            checkpoint['state_dict'] = OrderedDict((k[7:], v) for k, v in checkpoint['state_dict'].items())
        model.load_state_dict(checkpoint['state_dict'])
    else: 
        model.load_state_dict(checkpoint)

    if bi:
        frame_len4dataset = frame_len*2-1
    else:
        frame_len4dataset = frame_len
    test_dataset = VFSS_data(split_file = split_file, split = 'test', rgb_root=rgb_root, flow_root=flow_root, frame_len=frame_len4dataset, win_stride=win_stride,
        fill=fill, sampling=sampling, resize=resize, num_classes=num_classes, transforms=transform, front_rear_pad=front_rear_pad) # , gray_scale=gray_scale
    test_dataloader = torch.utils.data.DataLoader(test_dataset, batch_size=1, shuffle=False, pin_memory=False)    
    # print('dataloader length:',len(test_dataset))
    resultFile = f'{inferroot}/{saveName.split("_stride")[0]}.xlsx'
    wb = openpyxl.Workbook()
    sheet = wb[wb.active.title]
    sheet.append(['patientID', 'start', 'end', 'start_G', 'end_G', 'accuracy', 'f1', 'ap', 'infer_time', 'frame_len']) # interpolated 가져올때는 interpolation rate 곱해줘야 함

    # if (labelType != 'frameLabel') and not bi:
    #     ## VariTest_gradCAM.py 섞기#########################################################################
    #     from pytorch_grad_cam import GradCAM
    #     from pytorch_grad_cam.utils.image import show_cam_on_image
    #     # print('cannot import pytorch_grad_cam. pip install grad-cam')
    #     cam = GradCAM(model=model, target_layers=[target_layer], use_cuda=(device=="cuda")) # VariTest_gradCAM.py 섞기
    #     target_category = None # VariTest_gradCAM.py 섞기
    #     ## VariTest_gradCAM.py 섞기#########################################################################

    model.eval()
    y_true=[]
    y_score=[]
    for inputs, labels, iouLabel, patientID in tqdm(test_dataloader):
        start_time = timeit.default_timer()
        inputs=inputs[0] # (view_num, C, frame_len4dataset, H x W)
        iou_l = labels[0] # (num_classes, video_length)
        iou_l = iou_l[1] # (video_length)
        # print(f'video length: {len(iou_l)}')
        idx_l = list(range(len(iou_l))) 
        patientID = patientID.item()
        # print(inputs.shape, iou_l.shape)
        loop = inputs.size(0) // win_limit
        if inputs.size(0) % win_limit: loop = loop+1
        os.makedirs(f'{inferroot}/grad_cam/{patientID}', exist_ok=True)
        prob_l = None
        for i in range(loop):
            sub_input = inputs[i*win_limit:min(inputs.size(0), i*win_limit+win_limit)]
            if len(sub_input.size()) == 4: sub_input = sub_input.unsqueeze(0) # 배치가 1인경우 배치 생성
            if ('vgg' in modelName) and (frame_len == 1): sub_input = sub_input.squeeze(2)
            ###############################################################
            elif fill is not None:
                sub_input = sub_input.numpy()
                sub_input = np.transpose(sub_input, (2,1,0,3,4))
                sub_input = np.tile(sub_input, (tile_num, 1, 1, 1, 1))
                if partial_num != 0: sub_input=np.concatenate((sub_input, sub_input[:partial_num]))
                sub_input = np.transpose(sub_input, (2, 1, 0, 3, 4))
                sub_input = torch.from_numpy(sub_input)
            ##############################################################
            # print(f'sub input shape: {sub_input.shape}')
            sub_input = sub_input.to(device)
            # print(sub_input.size()) # (view_num, C, frame_len4dataset, H x W)
            with torch.no_grad(): # test니까 gradient안흐르게 model에 넣어보기. 아직 TDN 구현 안됨.
                if bi:# 0110, 00100
                    frame_prob = model(sub_input[:, :, :frame_len], sub_input[:, :, frame_len-1:])
                else:
                    frame_prob = model(sub_input)
                # print(frame_prob.size())
                if labelType == 'frameLabel':
                    frame_prob = torch.flatten(frame_prob, start_dim=0, end_dim=1)#torch.Size([(batch) x (frame_len), num_classes=2])
                    if front_rear_pad != 0:
                        frame_prob = frame_prob[:-front_rear_pad]
                if modelName=='TDN': frame_prob = frame_prob.squeeze(1)
                frame_prob = torch.nn.Softmax(dim=1)(frame_prob)# size: (stack_num, num_classes)
            frame_prob = torch.max(frame_prob, 1)[1].squeeze()
            # print(f'sub output shape: {frame_prob.shape}')
            # print(frame_prob.size())
            # ################################################################# VariTest_gradCAM.py 섞기
            # if (labelType != 'frameLabel') and not bi:
            #     # print(sub_input.size())#torch.Size([batch, RGB, frame_len, H, W])
            #     batch_size = sub_input.size(0)
            #     cam.batch_size = batch_size
            #     grayscale_cam = cam(input_tensor=sub_input, targets=target_category, aug_smooth=True, eigen_smooth=True)
            #     # print(grayscale_cam.shape)#torch.Size([batch, H, W, frame_len])
            #     grayscale_cam = torch.from_numpy(grayscale_cam)
            #     for b in range(batch_size):
            #         os.makedirs(f'{inferroot}/grad_cam/{patientID}/{i*batch_size+b+1}', exist_ok=True)
            #         # print(grayscale_cam[b].shape) # [H, W, frame_len]
            #         grayscale_cam_b = grayscale_cam[b].permute(2, 0, 1).contiguous()
            #         # print(grayscale_cam_b.shape) # [frame_len, H, W]
            #         sub_input_b = sub_input[b].permute(1, 2, 3, 0).contiguous()
            #         # print(sub_input_b.shape) #(frame_len, H, W, RGB)
            #         for f in range(sub_input_b.size(0)):
            #             grayscale_cam_bf = grayscale_cam_b[f]
            #             # print(grayscale_cam_bf.shape) # [H, W]
            #             gray_image = np.asarray(sub_input_b[f].cpu(), dtype=np.uint8) 
            #             # print(gray_image.shape) #(H, W, RGB)
            #             visualization = show_cam_on_image(gray_image, grayscale_cam_bf, use_rgb=True)
            #             visualization = Image.fromarray(visualization)
            #             visualization.save(f'{inferroot}/grad_cam/{patientID}/{patientID}_{i*batch_size+b+1}_{f}.png')
            #             visualization.save(f'{inferroot}/grad_cam/{patientID}/{i*batch_size+b+1}/{patientID}_{i*batch_size+b+1}_{f}.png')
            # ###########################################################################################

            # 모델에서 나온 결과값을 리스트에 저장
            if prob_l is None: 
                prob_l = frame_prob.tolist()
            elif type(frame_prob.tolist())==int:
                prob_l.append(frame_prob.item())
            else: prob_l.extend(frame_prob.tolist())

        prob_l = prob_l[:len(iou_l)] # cut off paddding part...used when labelType frameLabel

        plt.plot(idx_l, iou_l, idx_l, prob_l)
        plt.savefig(inferroot+f'/{patientID}.png')
        plt.clf()
        y_true.extend(iou_l)
        y_score.extend(prob_l)
        indices = [i for i, e in enumerate(prob_l) if e == 1]

        iou_l = torch.tensor(iou_l)
        prob_l = torch.tensor(prob_l)
        arg_labels = torch.argwhere(iou_l)
        accuracy = round(metrics.accuracy_score(iou_l, prob_l).item(), 4)
        F1_score = round(metrics.f1_score(iou_l, prob_l, pos_label=1, average='binary').item(), 4)
        ap = round(metrics.average_precision_score(iou_l, prob_l), 4)
        stop_time = timeit.default_timer()
        if indices != []: 
            sheet.append([patientID, indices[0], indices[-1], min(arg_labels).item(), max(arg_labels).item(), accuracy, F1_score, ap, str(stop_time - start_time), len(iou_l)])
        else:
            sheet.append([patientID, None, None, min(arg_labels).item(), max(arg_labels).item(), accuracy, F1_score, ap, str(stop_time - start_time), len(iou_l)])
        

    wb.save(resultFile)
    accuracy = metrics.accuracy_score(y_true, y_score) # corrects.double() / data_len
    recall = metrics.recall_score(y_true, y_score, pos_label=1, average='binary')# TP / (TP + FN)
    precision = metrics.precision_score(y_true, y_score, pos_label=1, average='binary')# TP / TPFP
    F1_score = metrics.f1_score(y_true, y_score, pos_label=1, average='binary')# 2*precision*recall/(precision+recall)
    fpr, tpr, thresholds = metrics.roc_curve(y_true, y_score, pos_label=1)
    accuracy = round(accuracy.item(), 3)
    recall = round(recall.item(), 3)
    precision = round(precision.item(), 3)
    F1_score = round(F1_score.item(), 3)
    auc = round(metrics.auc(fpr, tpr), 3)
    ap = round(metrics.average_precision_score(y_true, y_score), 3)
    print("Acc: {}".format(accuracy))
    print("recall: {}".format(recall))
    print("precision: {}".format(precision))
    print("F1_score: {}".format(F1_score))
    print("AUC: {}".format(auc))
    print("AP: {}".format(ap))
    plt.plot(fpr, tpr)
    plt.savefig(inferroot+f'/ROC_curve_testAcc-{accuracy}_recall-{recall}_precision-{precision}_F1score-{F1_score}_AUC-{auc}_AP-{ap}.png')
    plt.clf()
    print('the model was: ', last_model.split('/')[-1])

    resultFile = 'output.xlsx'
    if not os.path.isfile(os.path.join(save_dir, resultFile)):
        wb = openpyxl.Workbook()
        sheet = wb[wb.active.title]
        sheet.append(['LR', 'Test_acc', 'recall', 'precision', 'F1_score', 'AUC', 'AP', 'model'])
        wb.save(os.path.join(save_dir, resultFile))

    wb = openpyxl.load_workbook(resultFile)
    sheet = wb[wb.active.title] # Sheet1, Sheet
    sheet.append([LR, accuracy, recall, precision, F1_score, auc, ap, last_model.split('/')[-1][:-7]])
    wb.save(resultFile)

if __name__ == '__main__':
    import argparse
    # need to add argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--rgb_root', type=str, help='full path of the frames folder', default= None) #'frames'
    #server2:/home/DATA/syj/vfss/frames #local: C:/Users/singku/Downloads/vfss/frames #server4: /DATA/jsy/vfss/frames #lab: /home/DATA/syj/frames
    parser.add_argument('--flow_root', type=str, default=None, help='full path of the flows folder')
    parser.add_argument('--last_model', type=str, help='pth.tar file name', required=True)
    parser.add_argument('--split_file', type=str, help='split file name', default='bpm2ues_close_out.xlsx')
    parser.add_argument('--num_classes', type=int, help='number of action classes including the background', default = 2)
    # parser.add_argument('--win_stride', type=int, help='window sliding stride. at most frame_len', default=1)
    parser.add_argument('--win_limit', type=int, default=32, help='maximum frame batch that can be loaded on a GPU if it is smaller than the train batch size, win_limit is set to the train batch size')
    args = parser.parse_args()

    test_model(last_model=args.last_model, win_limit = args.win_limit, split_file=args.split_file, rgb_root = args.rgb_root, flow_root = args.flow_root, num_classes=args.num_classes) #win_stride=args.win_stride, 